"""
Export Module for SQL Agent
Supports multiple export formats: Excel, JSON, HTML, PDF.
"""

import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from typing import Optional
import base64


class DataExporter:
    """Handles exporting data to various formats."""
    
    def __init__(self):
        """Initialize the exporter."""
        pass
    
    def export_excel(self, df: pd.DataFrame, output_path: str,
                     sheet_name: str = "Query Results",
                     include_summary: bool = True) -> str:
        """
        Export DataFrame to Excel with formatting.
        
        Args:
            df: DataFrame to export
            output_path: Output file path
            sheet_name: Excel sheet name
            include_summary: Include summary statistics sheet
            
        Returns:
            str: Path to exported file
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Style header row
                header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add summary sheet if requested
                if include_summary:
                    numeric_cols = df.select_dtypes(include=['number']).columns
                    if len(numeric_cols) > 0:
                        summary_df = df[numeric_cols].describe()
                        summary_df.to_excel(writer, sheet_name="Summary Statistics")
            
            return output_path
            
        except ImportError:
            # Fallback to basic Excel export
            df.to_excel(output_path, sheet_name=sheet_name, index=False)
            return output_path
    
    def export_json(self, df: pd.DataFrame, output_path: str,
                    orient: str = 'records', indent: int = 2) -> str:
        """
        Export DataFrame to JSON.
        
        Args:
            df: DataFrame to export
            output_path: Output file path
            orient: JSON orientation ('records', 'index', 'columns')
            indent: JSON indentation
            
        Returns:
            str: Path to exported file
        """
        # Convert DataFrame to JSON
        json_data = df.to_json(orient=orient, date_format='iso')
        
        # Pretty print
        parsed = json.loads(json_data)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(parsed, f, indent=indent, ensure_ascii=False)
        
        return output_path
    
    def export_html(self, df: pd.DataFrame, output_path: str,
                    title: str = "Query Results",
                    include_search: bool = True) -> str:
        """
        Export DataFrame to interactive HTML.
        
        Args:
            df: DataFrame to export
            output_path: Output file path
            title: Page title
            include_search: Include search/filter functionality
            
        Returns:
            str: Path to exported file
        """
        # Generate HTML table
        table_html = df.to_html(index=False, classes='data-table', border=0)
        
        # Create full HTML page with styling
        html_template = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}
        
        .header p {{
            opacity: 0.9;
            font-size: 14px;
        }}
        
        .search-box {{
            padding: 20px 30px;
            background: #f8f9fa;
            border-bottom: 1px solid #dee2e6;
        }}
        
        .search-box input {{
            width: 100%;
            padding: 12px 20px;
            border: 2px solid #dee2e6;
            border-radius: 25px;
            font-size: 14px;
            outline: none;
            transition: all 0.3s;
        }}
        
        .search-box input:focus {{
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }}
        
        .table-container {{
            padding: 30px;
            overflow-x: auto;
        }}
        
        .data-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }}
        
        .data-table thead tr {{
            background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%);
            color: white;
        }}
        
        .data-table th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        
        .data-table td {{
            padding: 12px 15px;
            border-bottom: 1px solid #f0f0f0;
        }}
        
        .data-table tbody tr {{
            transition: background 0.2s;
        }}
        
        .data-table tbody tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        
        .data-table tbody tr:hover {{
            background: #e9ecef;
            transform: scale(1.01);
        }}
        
        .stats {{
            padding: 20px 30px;
            background: #f8f9fa;
            border-top: 1px solid #dee2e6;
            display: flex;
            justify-content: space-around;
            flex-wrap: wrap;
        }}
        
        .stat-item {{
            text-align: center;
            padding: 10px 20px;
        }}
        
        .stat-value {{
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
        }}
        
        .stat-label {{
            font-size: 12px;
            color: #6c757d;
            margin-top: 5px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{title}</h1>
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        {"<div class='search-box'><input type='text' id='searchInput' placeholder='🔍 Search table...' onkeyup='searchTable()'></div>" if include_search else ""}
        
        <div class="table-container">
            {table_html}
        </div>
        
        <div class="stats">
            <div class="stat-item">
                <div class="stat-value">{len(df):,}</div>
                <div class="stat-label">Total Rows</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{len(df.columns)}</div>
                <div class="stat-label">Columns</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{df.memory_usage(deep=True).sum() / 1024:.1f} KB</div>
                <div class="stat-label">Data Size</div>
            </div>
        </div>
    </div>
    
    {"<script>function searchTable() { var input = document.getElementById('searchInput'); var filter = input.value.toUpperCase(); var table = document.querySelector('.data-table'); var tr = table.getElementsByTagName('tr'); for (var i = 1; i < tr.length; i++) { var td = tr[i].getElementsByTagName('td'); var found = false; for (var j = 0; j < td.length; j++) { if (td[j]) { var txtValue = td[j].textContent || td[j].innerText; if (txtValue.toUpperCase().indexOf(filter) > -1) { found = true; break; } } } tr[i].style.display = found ? '' : 'none'; } }</script>" if include_search else ""}
</body>
</html>
"""
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_template)
        
        return output_path
    
    def export_pdf(self, df: pd.DataFrame, output_path: str,
                   title: str = "Query Results",
                   chart_path: Optional[str] = None) -> str:
        """
        Export DataFrame to PDF report.
        
        Args:
            df: DataFrame to export
            output_path: Output file path
            title: Report title
            chart_path: Optional path to chart image to include
            
        Returns:
            str: Path to exported file
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            # Create PDF
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=30,
                alignment=1  # Center
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 12))
            
            # Metadata
            meta_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Rows: {len(df):,} | Columns: {len(df.columns)}"
            elements.append(Paragraph(meta_text, styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Add chart if provided
            if chart_path and Path(chart_path).exists():
                img = Image(chart_path, width=6*inch, height=3*inch)
                elements.append(img)
                elements.append(Spacer(1, 20))
            
            # Prepare table data (limit to first 100 rows for PDF)
            table_data = [df.columns.tolist()]
            for idx, row in df.head(100).iterrows():
                table_data.append([str(val)[:50] for val in row.tolist()])  # Truncate long values
            
            # Create table
            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            
            elements.append(t)
            
            if len(df) > 100:
                elements.append(Spacer(1, 12))
                elements.append(Paragraph(f"<i>Showing first 100 of {len(df):,} rows</i>", styles['Italic']))
            
            # Build PDF
            doc.build(elements)
            return output_path
            
        except ImportError:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
